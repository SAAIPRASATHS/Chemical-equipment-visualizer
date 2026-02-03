"""
Data export utilities for generating Excel, CSV, and JSON exports.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from django.http import HttpResponse
from io import BytesIO
import json

from .models import Dataset, Equipment


class DataExporter:
    """Handles data export in multiple formats."""
    
    @staticmethod
    def export_to_excel(dataset_id):
        """
        Export dataset to Excel with formatting and charts.
        
        Args:
            dataset_id: ID of the dataset to export
            
        Returns:
            HttpResponse with Excel file
        """
        try:
            dataset = Dataset.objects.get(id=dataset_id)
            equipment_list = Equipment.objects.filter(dataset=dataset)
            
            # Create workbook
            wb = Workbook()
            
            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Add headers
            headers = ['Metric', 'Value']
            for col_num, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add summary data
            summary_data = [
                ['Dataset', dataset.filename],
                ['Upload Date', dataset.upload_date.strftime('%Y-%m-%d %H:%M')],
                ['Total Equipment', dataset.total_equipment],
                ['Healthy Count', dataset.healthy_count],
                ['Warning Count', dataset.warning_count],
                ['Critical Count', dataset.critical_count],
                ['Avg Flowrate', f"{dataset.avg_flowrate:.2f}"],
                ['Avg Pressure', f"{dataset.avg_pressure:.2f}"],
                ['Avg Temperature', f"{dataset.avg_temperature:.2f}°C"],
            ]
            
            for row_num, (metric, value) in enumerate(summary_data, 2):
                ws_summary.cell(row=row_num, column=1, value=metric).font = Font(bold=True)
                ws_summary.cell(row=row_num, column=2, value=value)
            
            # Equipment Details Sheet
            ws_equipment = wb.create_sheet("Equipment Details")
            
            # Headers
            eq_headers = ['Equipment Name', 'Type', 'Flowrate', 'Pressure', 'Temperature', 
                         'Health Score', 'Risk Level', 'Anomaly', 'Recommendations']
            for col_num, header in enumerate(eq_headers, 1):
                cell = ws_equipment.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Equipment data
            for row_num, eq in enumerate(equipment_list, 2):
                ws_equipment.cell(row=row_num, column=1, value=eq.name)
                ws_equipment.cell(row=row_num, column=2, value=eq.equipment_type)
                ws_equipment.cell(row=row_num, column=3, value=eq.flowrate)
                ws_equipment.cell(row=row_num, column=4, value=eq.pressure)
                ws_equipment.cell(row=row_num, column=5, value=eq.temperature)
                ws_equipment.cell(row=row_num, column=6, value=round(eq.health_score, 2))
                
                # Color code risk level
                risk_cell = ws_equipment.cell(row=row_num, column=7, value=eq.risk_level)
                if eq.risk_level == 'Critical':
                    risk_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    risk_cell.font = Font(color="C62828", bold=True)
                elif eq.risk_level == 'Warning':
                    risk_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                    risk_cell.font = Font(color="E65100", bold=True)
                else:
                    risk_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    risk_cell.font = Font(color="2E7D32", bold=True)
                
                ws_equipment.cell(row=row_num, column=8, value="Yes" if eq.is_anomaly else "No")
                ws_equipment.cell(row=row_num, column=9, value=eq.recommendations)
            
            # Adjust column widths
            for ws in [ws_summary, ws_equipment]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{dataset.filename}_export.xlsx"'
            
            return response
            
        except Dataset.DoesNotExist:
            return None
    
    @staticmethod
    def export_to_csv(dataset_id):
        """
        Export dataset to CSV format.
        
        Args:
            dataset_id: ID of the dataset to export
            
        Returns:
            HttpResponse with CSV file
        """
        try:
            dataset = Dataset.objects.get(id=dataset_id)
            equipment_list = Equipment.objects.filter(dataset=dataset)
            
            # Create DataFrame
            data = []
            for eq in equipment_list:
                data.append({
                    'Equipment Name': eq.name,
                    'Type': eq.equipment_type,
                    'Flowrate': eq.flowrate,
                    'Pressure': eq.pressure,
                    'Temperature': eq.temperature,
                    'Health Score': eq.health_score,
                    'Risk Level': eq.risk_level,
                    'Is Anomaly': eq.is_anomaly,
                    'Anomaly Score': eq.anomaly_score,
                    'Recommendations': eq.recommendations,
                })
            
            df = pd.DataFrame(data)
            
            # Create response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{dataset.filename}_export.csv"'
            
            df.to_csv(response, index=False)
            
            return response
            
        except Dataset.DoesNotExist:
            return None
    
    @staticmethod
    def export_to_json(dataset_id):
        """
        Export dataset to JSON format.
        
        Args:
            dataset_id: ID of the dataset to export
            
        Returns:
            HttpResponse with JSON file
        """
        try:
            dataset = Dataset.objects.get(id=dataset_id)
            equipment_list = Equipment.objects.filter(dataset=dataset)
            
            # Create export data
            export_data = {
                'dataset': {
                    'filename': dataset.filename,
                    'upload_date': dataset.upload_date.isoformat(),
                    'total_equipment': dataset.total_equipment,
                    'summary': {
                        'healthy_count': dataset.healthy_count,
                        'warning_count': dataset.warning_count,
                        'critical_count': dataset.critical_count,
                        'avg_flowrate': dataset.avg_flowrate,
                        'avg_pressure': dataset.avg_pressure,
                        'avg_temperature': dataset.avg_temperature,
                    },
                    'executive_summary': dataset.executive_summary,
                },
                'equipment': []
            }
            
            for eq in equipment_list:
                export_data['equipment'].append({
                    'name': eq.name,
                    'type': eq.equipment_type,
                    'flowrate': eq.flowrate,
                    'pressure': eq.pressure,
                    'temperature': eq.temperature,
                    'health_score': eq.health_score,
                    'risk_level': eq.risk_level,
                    'is_anomaly': eq.is_anomaly,
                    'anomaly_score': eq.anomaly_score,
                    'anomaly_reasons': eq.anomaly_reasons,
                    'recommendations': eq.recommendations,
                })
            
            # Create response
            response = HttpResponse(
                json.dumps(export_data, indent=2),
                content_type='application/json'
            )
            response['Content-Disposition'] = f'attachment; filename="{dataset.filename}_export.json"'
            
            return response
            
        except Dataset.DoesNotExist:
            return None
